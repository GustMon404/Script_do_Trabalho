from datetime import date,datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
import os
os.chdir('C:\\Users\\USER\\Desktop\\RH\\Escala')

escala = load_workbook('escala_copy.xlsx')
folha = load_workbook('Folha_de_Ponto_Original.xlsx')
sheet= escala['Plan1']
nome_folha = folha.sheetnames
semana = ['Seg','Ter','Qua','Qui','Sex','Sab','Dom']

ano_atual = date.today().year
mes_atual = date.today().month

for nomes in nome_folha:
    sheet1 = folha[nomes]
    for i in range(1,32):        
        #A exceção é por conta dos meses não ter a mesma quantidade de dias, gera erro o for contar até 31 sendo que tem mes que vai até 28        
        try:
            data_mes = date(ano_atual, mes_atual, i).isocalendar() #Exibe o dia da semana em numeros Ex: segunda: 1 e domingo: 7, 2020/01/01 é Quarta
            dia = semana[data_mes[2]-1]# segunda é 1, porem na lista semana, é 0
            sheet1['B'+str(i+13)].value = i
            sheet1['C'+str(i+13)].value = dia
        except:
            break

for i in range (1,18):
    coluna_funcionario = sheet.cell(row=4, column=i+4).value
    if coluna_funcionario != None:
        planilha_folha = folha[coluna_funcionario]
        for a in range(1,32):
            folga = sheet.cell(row = 4+a, column = i+4).value
            if folga == 'FOLGA':
                planilha_folha['D'+str(13+a)].value = 'Folga'
                planilha_folha['E'+str(13+a)].value = '***'
                planilha_folha['F'+str(13+a)].value = '***'
                planilha_folha['G'+str(13+a)].value = '***'
                planilha_folha['H'+str(13+a)].value = '***'
                
#Define o tamanho e cor da borda de todos os lados que minha celula usara futuramente
thin_border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

fonte = Font(name='Arial', size = 12, bold = True, italic = False)

alinhamento = Alignment(horizontal='center')

#não achei nem uma função em datetime que retorna o maximo de dias de um mes
'''for linha in range (1,32):
    try:
        date(ano_atual,mes_atual, linha)
        for a in range (1,10):
            for nomes in range(len(nome_folha)):
                folha[nome_folha[nomes]].cell(row=linha+13, column=a+1).border = thin_border
    except:
        break'''
for linha in range (1,32):
    try:
        date(ano_atual,mes_atual, linha)
        for nomes in range(len(nome_folha)):
            for coluna in range (1,10):
                folha[nome_folha[nomes]].cell(row=linha+13, column=coluna+1).border = thin_border
                folha[nome_folha[nomes]].cell(row=linha+13, column=coluna+1).font = fonte
                folha[nome_folha[nomes]].cell(row=linha+13, column=coluna+1).alignment = alinhamento
    except:
        break
folha.save('folha_copy.xlsx')
