from datetime import date
from openpyxl import load_workbook
import os
os.chdir('C:\\Users\\USER\\Desktop\\RH\\Escala')

funB = {'Daniel':'seg',  'Cida':'dom','Natanael':'qui', 'Itaci':'sab', 'Raphael':'sab',
       'Gildasio':'sex', 'Aline':'sab'}
funP = {'Adriana':'seg', 'Sand':'ter', 'Andreza':'qua', 'Paula':'qui', 'Fernanda':'sex',
        'Lucimaria':'qua'}
plant = {'Pollyana': 'domingo','Rita':'dom'}

folga_domingo = {'Daniel':1,  'Cida':1,'Natanael':1, 'Itaci':1, 'Raphael':1,
       'Gildasio':2, 'Aline':3}
#wb = Workbook()
ws = load_workbook('escala_copy.xlsx')
sheet= ws['Plan1']  
semana = ['seg','ter','qua','qui','sex','sab','dom']


for i in range(1,32):
    ano_atual = date.today().year
    mes_atual = date.today().month
    try:
        data_mes = date(ano_atual, mes_atual, i).isocalendar() #Exibe o dia da semana em numeros Ex: segunda: 1 e domingo: 7
        dia = semana[data_mes[2]-1] # segunda é 1, porem na lista semana, é 0
        sheet['D'+str(i+4)].value= sheet['O'+str(i+4)].value= dia
        sheet['C'+str(i+4)].value= sheet['N'+str(i+4)].value= i
    except:
        break


listaB = list(funB.keys())

for i in range(len(funB)):
    #vai adicionar o nome na coluna
    linha_nome = sheet.cell(row=4, column=i+5).value = str(listaB[i])
    for dia in range(1,32):
        coluna_balconista = sheet.cell(row=dia+4, column=i+5)
        if sheet['D'+str(dia+4)].value == funB[linha_nome]:
            coluna_balconista.value = 'FOLGA'
        # Os vendedores trabalham 2 domingo para folgar 1, foi criado uma dicionario,
        #com a numeração do ultimo domingo que é igual a 1
        if sheet['D'+str(dia+4)].value == 'dom':
            #quando for 3 a numeração do vendedor ele poderá folgar
            if folga_domingo[linha_nome] == 3:
                coluna_balconista.value = 'FOLGA'
                folga_domingo[linha_nome] = 1
            else:
                folga_domingo[linha_nome] += 1

                
#Não queria ter q usar a mesma função para fazer algo parecido
inicio_colunaP = len(funB) + 9
listaP = list(funP.keys())
for i in range(len(funP)):
    linha_nome = sheet.cell(row=4, column=i+inicio_colunaP).value = str(listaP[i])
    if linha_nome in funP:
        for dia in range(1,32):
            coluna_balconista = sheet.cell(row=dia+4, column=i+inicio_colunaP)
            if sheet['O'+str(dia+4)].value == funP[linha_nome]:
                coluna_balconista.value = 'FOLGA'
            
'''
#refiz os dois laços, que preenchia a tabela, porem, ela se baseia em uma tabela ja preenchida com os nomes
qt_fun=1
a=0
funB.update(funP)
while True:
    nome_funcionario = sheet.cell(row=4, column=qt_fun).value
    if nome_funcionario in funB or nome_funcionario in funP:
        for dia in range(1,32):
            coluna_funcionario = sheet.cell(row=dia+4, column=qt_fun)
            if sheet['D'+str(dia+4)].value == funB[nome_funcionario]:
                coluna_funcionario.value='FOLGA'
        a+=1
    if len(funB)==a:
        break
    qt_fun +=1
'''    
ws.save('escala_copy.xlsx')
cmd = r"C:\Users\USER\Desktop\RH\Escala\escala_copy.xlsx"
os.system(cmd)
